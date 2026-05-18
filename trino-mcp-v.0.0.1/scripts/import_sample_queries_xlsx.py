#!/usr/bin/env python3
"""Convert REN3_SAMPLE_QUESTIONS_QUERIES.xlsx → samples/sample_queries.yaml.

Usage:
  uv run python scripts/import_sample_queries_xlsx.py /path/to/REN3_SAMPLE_QUESTIONS_QUERIES.xlsx
  uv run python scripts/import_sample_queries_xlsx.py /path/to/file.xlsx -o samples/sample_queries.yaml
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import yaml

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required: uv sync --extra dev", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT = ROOT / "samples" / "sample_queries.yaml"

QUESTION_HEADERS = (
    "ren3 prompt",
    "question",
    "sample question",
    "questions",
    "natural language",
    "nl question",
    "user question",
    "business question",
    "prompt",
    "query (natural language)",
)
SQL_HEADERS = (
    "sql",
    "sql query",
    "sql translation",
    "translated sql",
    "trino sql",
)
# Avoid matching "query" as substring of "ren3 prompt" — exact header match only for sql column.
SQL_HEADERS_FUZZY = ("sql", "sql query", "sql translation", "translated sql", "trino sql")
CATEGORY_HEADERS = ("category", "topic", "section", "domain", "project")
TAGS_HEADERS = ("tags", "tag")
NOTES_HEADERS = ("notes", "note", "comments", "comment", "remarks", "description")


def _norm_header(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _pick_column(
    headers: list[str],
    candidates: tuple[str, ...],
    *,
    fuzzy: tuple[str, ...] | None = None,
) -> int | None:
    for i, h in enumerate(headers):
        if h in candidates:
            return i
    if fuzzy:
        for i, h in enumerate(headers):
            if any(c in h for c in fuzzy if len(c) > 3):
                return i
    return None


def _slug_id(text: str, index: int) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")[:48]
    return f"q{index:03d}_{slug}" if slug else f"q{index:03d}"


def _cell_str(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def import_workbook(xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    all_queries: list[dict] = []
    global_index = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            continue

        headers = [_norm_header(c) for c in header_row]
        q_col = _pick_column(headers, QUESTION_HEADERS)
        s_col = _pick_column(headers, SQL_HEADERS, fuzzy=SQL_HEADERS_FUZZY)
        c_col = _pick_column(headers, CATEGORY_HEADERS)
        t_col = _pick_column(headers, TAGS_HEADERS)
        n_col = _pick_column(headers, NOTES_HEADERS)

        if q_col is None and len(headers) >= 3:
            # Last resort: column before SQL is often the NL prompt
            if s_col is not None and s_col > 0:
                q_col = s_col - 1
        if q_col is None or s_col is None:
            if len(headers) >= 2 and headers[0] and headers[1]:
                if q_col is None:
                    q_col = 0
                if s_col is None:
                    s_col = 1
            else:
                print(
                    f"Skipping sheet {sheet_name!r}: no question/sql columns "
                    f"(headers={headers!r})",
                    file=sys.stderr,
                )
                continue
        print(
            f"Sheet {sheet_name!r}: question={headers[q_col]!r} sql={headers[s_col]!r}"
            + (f" category={headers[c_col]!r}" if c_col is not None else ""),
            file=sys.stderr,
        )

        for row in rows:
            if row is None:
                continue
            cells = list(row)
            if q_col >= len(cells):
                continue
            question = _cell_str(cells[q_col])
            sql = _cell_str(cells[s_col]) if s_col < len(cells) else ""
            if not question or not sql:
                continue

            global_index += 1
            category = _cell_str(cells[c_col]) if c_col is not None and c_col < len(cells) else ""
            if not category and sheet_name.lower() not in ("sheet1", "sheet"):
                category = sheet_name.strip()

            tags_raw = _cell_str(cells[t_col]) if t_col is not None and t_col < len(cells) else ""
            tags = [t.strip() for t in re.split(r"[,;]", tags_raw) if t.strip()]
            notes = _cell_str(cells[n_col]) if n_col is not None and n_col < len(cells) else ""

            all_queries.append(
                {
                    "id": _slug_id(question, global_index),
                    "question": question,
                    "sql": sql,
                    "category": category,
                    "tags": tags,
                    "notes": notes,
                }
            )

    wb.close()
    return {
        "version": 1,
        "source": xlsx_path.name,
        "queries": all_queries,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "xlsx",
        type=Path,
        nargs="?",
        default=ROOT / "samples" / "REN3_SAMPLE_QUESTIONS_QUERIES.xlsx",
        help="Path to REN3_SAMPLE_QUESTIONS_QUERIES.xlsx (default: samples/ in repo)",
    )
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUT)
    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"File not found: {args.xlsx}", file=sys.stderr)
        sys.exit(1)

    data = import_workbook(args.xlsx)
    if not data["queries"]:
        print("No rows imported — check sheet headers (question + sql columns).", file=sys.stderr)
        sys.exit(1)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        yaml.safe_dump(data, sort_keys=False, allow_unicode=True, width=1000),
        encoding="utf-8",
    )
    print(f"Wrote {len(data['queries'])} queries to {args.output}")


if __name__ == "__main__":
    main()
